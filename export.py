from pprint import pprint
import random
import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side

column = ["A", "B", "C", "D", "E"]
categoryDict = {
    "cvrcek": "Cvrček",
    "benjamin": "Benamín",
    "junior": "Junior",
    "kadet": "Kadet",
    "klokanek": "Klokánek",
    "student": "Student",
}

columnDictionary = {
    "score": "A",
    "name": "B",
    "surname": "C",
    "grade": "D",
    "birthdate": "E",
    "school": "F",
}

minWidth = {"A": 6, "B": 7, "C": 10, "D": 7, "E": 17, "F": 7}

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

class file:
    def __str__(self) -> str:
        return "export/" + self.name + ".xlsx"


class userExport(file):
    def __init__(self, results, name) -> None:
        self.workspace = excel.load_workbook("template.xlsx")
        self.results = results
        self.name = name
        self.fill()
        self.workspace.save("export/" + self.name + ".xlsx")

    def fill(self) -> None:
        for category in self.results["score"]:
            templateSheet = self.workspace.active
            sheet = self.workspace.copy_worksheet(templateSheet)
            sheet.title = categoryDict[category]

            # filling template
            sheet["B2"] = self.results["school"]
            sheet["B3"] = self.results["address"]
            sheet["B4"] = categoryDict[category]

            # filling points
            score = len(self.results["score"][category])
            for i in range(score - 1, -1, -1):
                sheet["A" + str(score - i + 8)] = i
                sheet["B" + str(score - i + 8)] = self.results["score"][category][
                    str(i)
                ]

            # filling solvers
            for index, x in enumerate(self.results["best"][category]):
                solver = self.results["best"][category][x]
                for dataIndex, dataPoint in enumerate(solver):
                    sheet[column[dataIndex] + str(index + 9)] = solver[dataPoint]

        del self.workspace["Sheet1"]


class finalExport(file):
    def __init__(self, results) -> None:
        self.workspace = excel.load_workbook("final.xlsx")
        self.template = self.workspace.active
        self.name = "MatKlokan - 2022"
        self.results = results
        self.columnWidth = {}
        self.fill()
        self.workspace.save("export/" + self.name + ".xlsx")

    def parseWidth(self, text, column):
        if column in self.columnWidth:
            self.columnWidth[column] = max(
                [self.columnWidth[column], len(str(text)), 5]
            )
        else:
            try:
                self.columnWidth[column] = max([len(str(text)), minWidth[column]])
            except:
                self.columnWidth[column] = max([len(str(text)), 5])

    def fill(self) -> None:
        for key,category in self.results.items():
            self.columnWidth = {}
            maxScore = int(category["range"]["biggest"])
            rowValue = [0 for _ in range(maxScore)]
            
            sheet = self.workspace.copy_worksheet(self.template)
            sheet.title = categoryDict[key]

            finalSum = 0

            for row in range(maxScore, -1, -1):
                coord = "M" + str(maxScore - row + 2) 
                sheet[coord] = row
                sheet[coord].fill = PatternFill(
                    start_color="d9d9d9", patternType="solid"
                )
                sheet.cell(row=maxScore - row + 2, column=13).border = thin_border

            # school score fill
            for index, school in enumerate(category["score"]):
                cell = sheet.cell(row=1, column=index + 14)
                cell.value = school
                cell.border = thin_border
                cell.fill = PatternFill(
                    start_color="a9d08e", patternType="solid"
                )

                for row in range(maxScore, -1, -1):
                    try:
                        value = category["score"][school][str(row)]
                        sheet.cell(row=maxScore - row + 2, column=index + 14).value = value
                        rowValue[maxScore - row] += value
                        finalSum += value
                    except:
                        pass
            
            for row, value in enumerate(rowValue):
                sheet.cell(row=row+2,column=12).value = value
                sheet.cell(row=row+2,column=12).border = thin_border
            sheet["J2"] = finalSum

            # solver fill - done
            for index, solver in enumerate(category["best"]):
                for datapoint in category["best"][solver]:
                    try:
                        value = category["best"][solver][datapoint]
                        sheet[columnDictionary[datapoint] + str(index + 3)] = value
                        self.parseWidth(value, columnDictionary[datapoint])
                    except:
                        pass
            
            #column resize
            for each in self.columnWidth:
                sheet.column_dimensions[each].width = self.columnWidth[each]

        del self.workspace["Sheet1"]
